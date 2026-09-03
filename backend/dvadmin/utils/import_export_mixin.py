# -*- coding: utf-8 -*-
from types import FunctionType, MethodType
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request

from dvadmin.utils.import_export import import_to_data
from dvadmin.utils.json_response import DetailResponse
from dvadmin.utils.request_util import get_verbose_name


class ImportSerializerMixin:
    """
    Fonctionnalités personnalisées de modèle d'import et d'import
    """

    # Champs d'import
    import_field_dict = {}
    # Sérialiseur d'import
    import_serializer_class = None
    # Largeur maximale de l'en-tête du tableau, 50 caractères par défaut
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        Obtenir la longueur maximale d'une chaîne de caractères
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get','post'],detail=False)
    @transaction.atomic  # Transaction Django pour éviter les erreurs
    def import_data(self, request: Request, *args, **kwargs):
        """
        Modèle d'import
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        assert self.import_field_dict, "'%s' Veuillez configurer les champs du modèle d'export correspondants." % self.__class__.__name__
        if isinstance(self.import_field_dict, MethodType) or isinstance(self.import_field_dict, FunctionType):
            self.import_field_dict = self.import_field_dict()
        # Exporter le modèle
        if request.method == "GET":
            # Données d'exemple
            queryset = self.filter_queryset(self.get_queryset())
            # Exporter le tableau Excel
            response = HttpResponse(content_type="application/msexcel")
            response["Access-Control-Expose-Headers"] = f"Content-Disposition"
            response[
                "Content-Disposition"
            ] = f'attachment;filename={quote(str(f"Importer le modèle {get_verbose_name(queryset)}.xlsx"))}'
            wb = Workbook()
            ws1 = wb.create_sheet("data", 1)
            ws1.sheet_state = "hidden"
            ws = wb.active
            row = get_column_letter(len(self.import_field_dict) + 1)
            column = 10
            header_data = [
                "N°",
            ]
            validation_data_dict = {}
            for index, ele in enumerate(self.import_field_dict.values()):
                if isinstance(ele, dict):
                    header_data.append(ele.get("title"))
                    choices = ele.get("choices", {})
                    if choices.get("data"):
                        data_list = []
                        data_list.extend(choices.get("data").keys())
                        validation_data_dict[ele.get("title")] = data_list
                    elif choices.get("queryset") and choices.get("values_name"):
                        data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                        validation_data_dict[ele.get("title")] = list(data_list)
                    else:
                        continue
                    column_letter = get_column_letter(len(validation_data_dict))
                    dv = DataValidation(
                        type="list",
                        formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[ele.get('title')]) + 1}",
                        allow_blank=True,
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{get_column_letter(index + 2)}2:{get_column_letter(index + 2)}1048576")
                else:
                    header_data.append(ele)
            # Ajouter les colonnes de données
            ws1.append(list(validation_data_dict.keys()))
            for index, validation_data in enumerate(validation_data_dict.values()):
                for inx, ele in enumerate(validation_data):
                    ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
            # Insérer les données officielles du modèle d'export
            df_len_max = [self.get_string_len(ele) for ele in header_data]
            ws.append(header_data)
            # Mettre à jour la largeur des colonnes
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            tab = Table(displayName="Table1", ref=f"A1:{row}{column}")  # Gestionnaire de noms
            style = TableStyleInfo(
                name="TableStyleLight11",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(response)
            return response
        else:
            # Organiser la structure de données correspondante depuis Excel, puis l'enregistrer avec le sérialiseur
            queryset = self.filter_queryset(self.get_queryset())
            # Récupérer les champs plusieurs-à-plusieurs
            m2m_fields = [
                ele.name
                for ele in queryset.model._meta.get_fields()
                if hasattr(ele, "many_to_many") and ele.many_to_many == True
            ]
            import_field_dict = {'id':'Update key (do not modify)',**self.import_field_dict}
            data = import_to_data(request.data.get("url"), import_field_dict, m2m_fields)
            for ele in data:
                filter_dic = {'id':ele.get('id')}
                instance = filter_dic and queryset.filter(**filter_dic).first()
                # print(156,ele)
                serializer = self.import_serializer_class(instance, data=ele, request=request)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            return DetailResponse(msg=f"Import réussi !")

    @action(methods=['get'],detail=False)
    def update_template(self,request):
        queryset = self.filter_queryset(self.get_queryset())
        assert self.import_field_dict, "'%s' Veuillez configurer les champs du modèle d'import correspondants." % self.__class__.__name__
        assert self.import_serializer_class, "'%s' Veuillez configurer le sérialiseur d'import correspondant." % self.__class__.__name__
        data = self.import_serializer_class(queryset, many=True, request=request).data
        if isinstance(self.import_field_dict, MethodType) or isinstance(self.import_field_dict, FunctionType):
            self.import_field_dict = self.import_field_dict()
        # Exporter le tableau Excel
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"Exporter {get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active
        import_field_dict = {}
        header_data = ["N°","Update key (do not modify)"]
        hidden_header = ["#","id"]
        #----Configurer les options----
        validation_data_dict = {}
        for index, item in enumerate(self.import_field_dict.items()):
            items = list(item)
            key = items[0]
            value = items[1]
            if isinstance(value, dict):
                header_data.append(value.get("title"))
                hidden_header.append(value.get('display'))
                choices = value.get("choices", {})
                if choices.get("data"):
                    data_list = []
                    data_list.extend(choices.get("data").keys())
                    validation_data_dict[value.get("title")] = data_list
                elif choices.get("queryset") and choices.get("values_name"):
                    data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                    validation_data_dict[value.get("title")] = list(data_list)
                else:
                    continue
                column_letter = get_column_letter(len(validation_data_dict))
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[value.get('title')]) + 1}",
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 3)}2:{get_column_letter(index + 3)}1048576")
            else:
                header_data.append(value)
                hidden_header.append(key)
        # Ajouter les colonnes de données
        ws1.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
        #--------
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(hidden_header) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if val is None or val == "":
                            results_list.append("")
                        elif isinstance(val,list):
                            results_list.append(str(val))
                        else:
                            results_list.append(val)
                        # Calculer la largeur maximale des colonnes
                        if isinstance(val,str):
                            result_column_width = self.get_string_len(val)
                            if h_index != 0 and result_column_width > df_len_max[h_index]:
                                df_len_max[h_index] = result_column_width
            ws.append([index+1,*results_list])
            column += 1
        # Mettre à jour la largeur des colonnes
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # Gestionnaire de noms
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response


class ExportSerializerMixin:
    """
    Fonctionnalité d'export personnalisée
    """

    # Champs d'export
    export_field_label = []
    # Sérialiseur d'export
    export_serializer_class = None
    # Largeur maximale de l'en-tête du tableau, 50 caractères par défaut
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        Obtenir la longueur maximale d'une chaîne de caractères
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get'],detail=False)
    def export_data(self, request: Request, *args, **kwargs):
        """
        Fonctionnalité d'export
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        queryset = self.filter_queryset(self.get_queryset())
        assert self.export_field_label, "'%s' Veuillez configurer les champs du modèle d'export correspondants." % self.__class__.__name__
        assert self.export_serializer_class, "'%s' Veuillez configurer le sérialiseur d'export correspondant." % self.__class__.__name__
        data = self.export_serializer_class(queryset, many=True, request=request).data
        # Exporter le tableau Excel
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"Exporter {get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws = wb.active
        header_data = ["N°", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key,val in results.items():
                    if key == h_item:
                        if val is None or val=="":
                            results_list.append("")
                        else:
                            results_list.append(val)
                        # Calculer la largeur maximale des colonnes
                        result_column_width = self.get_string_len(val)
                        if h_index !=0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        # Mettre à jour la largeur des colonnes
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # Gestionnaire de noms
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response
